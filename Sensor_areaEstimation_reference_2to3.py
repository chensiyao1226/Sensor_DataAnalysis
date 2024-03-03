import numpy as np
import pandas as pd
import os
import openpyxl as xl
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import time
from scipy.optimize import curve_fit

'''
def func(x,t0,c0,c1,k):
    return c0 + (c1 - c0) * np.exp(-k*(x-t0))
'''
def funcE(x,c0,c1,k):
    return c0 + c1 * np.exp(-k*x)
def funcL(x,c0,k):
    return c0 + k*x

def fitting():
    t = np.linspace(1, len(temp_predict_dat)+1, len(temp_predict_dat))
    y = temp_predict_dat
    lenY = len(temp_predict_dat)
    plt.plot(t*10, y, 'bo')

    temp_gap = abs(np.mean(temp_predict_dat[0:1]) - np.mean(temp_predict_dat[lenY - 2:lenY - 1]))
    print("temp_gap")
    print(temp_gap)
    if(temp_gap > 2):
        flag=0
        popt, pcov = curve_fit(funcE, t, y)
    else:
        flag=1
        popt, pcov = curve_fit(funcL, t, y)

    t2 = np.linspace(0, len(temp_predict_dat)+1, 10*len(temp_predict_dat))#0.4
    if (flag==0):
        y2 = [funcE(i, popt[0], popt[1], popt[2]) for i in t2]
    else:
        y2 = [funcL(i, popt[0], popt[1]) for i in t2]

    plt.plot(t2*10, y2, 'r--')
    print(popt)

    if (flag == 0):
        predic_temp = funcE(0, popt[0], popt[1], popt[2])#0.5
    else:
        predic_temp = funcL(0, popt[0], popt[1])#0.5
    plt.plot(0, predic_temp, 'ro')#0

    if (flag==0):
        residuals = y - funcE(t, *popt)
    else:
        residuals = y - funcL(t, *popt)

    ss_res = np.sum(residuals ** 2)
    ss_tot = np.sum((y - np.mean(y)) ** 2)
    r_squared = 1 - (ss_res / ss_tot)
    print("r_squared")
    print(r_squared)

    ax = plt.gca()
    plt.tick_params(labelsize=17)
    font2 = {'size': 22}
    ax.set_xlabel("Time (s)", font2)
    ax.set_ylabel("Temperature (℃)", font2)

    plt.show()
    return flag, popt, r_squared




path = "C:\\Users\chen_\OneDrive - Kyoto University\デスクトップ\研究\chen cell\\2. Food inspection\Flow strategy\\2.27"
samp_filename = input('input filename:')
samp_file_magn = (path + '\\' + samp_filename + '.csv')

samp_sheet = pd.read_csv(samp_file_magn, skiprows=5)
start_moment = int(input('input start moment:'))
samp_moment = int(input('input sample moment:'))
#ref_moment = int(input('input reference moment:'))
tempEnd_moment = int(input('input temperature terminus:'))

samp_dat = samp_sheet.iloc[samp_moment,8:]
samp_dat = 1000*samp_dat.values


ref_dat_1 = samp_sheet.iloc[samp_moment+2,8:]
ref_dat_1 = 1000*ref_dat_1.values
temp_ref_1 = np.mean(samp_sheet.values[samp_moment+2,2:6])

ref_dat_2 = samp_sheet.iloc[samp_moment+3,8:]
ref_dat_2 = 1000*ref_dat_2.values
temp_ref_2 = np.mean(samp_sheet.values[samp_moment+3,2:6])



ref_dat = (ref_dat_1 + ref_dat_2)/2
temp_ref = (temp_ref_1 + temp_ref_2)/2

STD_ref = samp_sheet.iloc[samp_moment+10,8:]
STD_ref = 1000*STD_ref.values

print("ref_dat")
print(ref_dat)
print("temp_ref")
print(temp_ref)

temp_predict_dat = samp_sheet.iloc[(samp_moment+1):tempEnd_moment,2:6]
temp_predict_dat = np.mean(temp_predict_dat.values, axis=1)


##seg = air - (ref + cal.mean(axis=0))/2
##samp_shift = samp_dat - ref_samp_dat

##===========温度予測==================
flagL, parameter, r_square = fitting()

if(flagL==0):
    samp_temp = funcE(0.5, parameter[0], parameter[1], parameter[2])
else:
    samp_temp = funcL(0.5, parameter[0], parameter[1])
print(samp_temp)

#=======respective============
path_ref_cali = "C:\\Users\chen_\OneDrive - Kyoto University\デスクトップ\研究\chen cell\cb line\\samp Line respective 1.xlsx"
cali_excel = xl.load_workbook(path_ref_cali)
cali_sheet = cali_excel.active
t_alp = np.zeros(1488)
b_alp = np.zeros(1488)
s_alp = np.zeros(1488)

for i in range (0,1488):
    t_alp[i] = cali_sheet.cell(row=1,column=i+1).value
    b_alp[i] = cali_sheet.cell(row=2,column=i+1).value
    s_alp[i] = cali_sheet.cell(row=3,column=i+1).value


t_array = np.zeros((1488,1488))
b_array = np.zeros((1488,1488))
s_array = np.zeros((1488,1488))
for i in range(0, 1488):
    t_array[i, i] = t_alp[i]
    b_array[i, i] = b_alp[i]
    s_array[i, i] = s_alp[i]

#======= shift calculation==========
temp_gap = samp_temp-temp_ref
temp_compen = np.dot(temp_gap,t_alp)
temp_compen_mean = np.mean(temp_compen)

ref_dat_modify = ref_dat+temp_compen

shift_comp = samp_dat - ref_dat_modify
shift_comp_modi = np.dot(shift_comp, s_array)
shift_comp_modi_matrix = np.reshape(shift_comp_modi, (62, 24))

shift_noComp = samp_dat - ref_dat
shift_noComp_modi = np.dot(shift_noComp, s_array)
shift_noComp_modi_matrix = np.reshape(shift_noComp_modi, (62, 24))

#========data overtime===========
STD_ref = STD_ref.astype(np.float64)

samp_overtime = samp_sheet.iloc[start_moment:tempEnd_moment,8:]
samp_overtime = 1000*samp_overtime.values

shift_overtime = samp_overtime - STD_ref ##以最后sample时刻的温度为准



shift_modi_overtime = np.zeros((len(samp_overtime), 1488))

for i in range (0,len(samp_overtime)):
    shift_modi_overtime[i] = np.dot(shift_overtime[i], s_array)


samp_STD = np.zeros(len(samp_overtime))
samp_ave = np.zeros(len(samp_overtime))
for i in range(0, len(samp_overtime)):
    samp_STD[i] = np.std(shift_modi_overtime[i],ddof=1)
    samp_ave[i] = np.mean(samp_overtime[i])

samp_STD_list = samp_STD.tolist()
samp_ave_list = samp_ave.tolist()

#========mean shift calculation================
#====area=====
a1, a3, a2, a4 = map(int, (0,0,63,25))
shift_comp_area_mean = np.mean(shift_comp_modi_matrix[a1:a2,a3:a4])
shift_noComp_area_mean = np.mean(shift_noComp_modi_matrix[a1:a2,a3:a4])

#==unmodify===
samp_shift_matrix = np.reshape(shift_comp, (62, 24))
samp_area_mean_unmodify = np.mean(samp_shift_matrix[a1:a2,a3:a4])

print("shift_comp_area_mean")
print(shift_comp_area_mean)

print("shift_noComp_area_mean")
print(shift_noComp_area_mean)
#========show image===============
###領域設定====
region = np.empty([2,62,24])
h = np.linspace(0,61,62)
v = np.linspace(0,23,24)
v,h = np.meshgrid(v,h)
region[0,:,:] = 62-h
region[1,:,:] = ((24-v)*2-(region[0,:,:]%2))/2


region_analyse = region[:,a1:a2,a3:a4]
h_val = 110
v_val = 50
vert = [(-h_val,-v_val),(h_val,-v_val),(h_val,v_val),(-h_val,v_val)]
vert = np.array(vert)

fig, ax = plt.subplots()
ax.invert_xaxis()
ax.set_xlabel('v', fontsize=14)
ax.set_ylabel('h', fontsize=14)
mappable = ax.scatter(region[1,:,:], region[0,:,:], c=shift_noComp_modi_matrix, cmap=cm.jet_r, marker=vert)

#vmin=-5, vmax=48
cbar=plt.colorbar(mappable)
cbar.ax.tick_params(labelsize=13)
plt.xticks(fontsize=14)
plt.yticks(fontsize=14)

#fig.colorbar(mappable)
fig.tight_layout()

plt.gcf().subplots_adjust(bottom=0.15)
#plt.savefig("C:\\Users\dell\Desktop\研究\chen cell\\retake\\EcoliCMOS.pdf")
#plt.fill_between((a3, a4-1), a1, a2-1, facecolor='red', alpha=0.2)
plt.show()

#========save===========
##path = "C:\\Users\dell\Desktop\研究\chen cell\calibration line\CMOS\sample\91"
print('Input Excel Name')
Excel_name = "DEP spectrum Near23 cali 1"##input('>>')
DEP_f = input("DEP Frequency/ Data Title:")
print("creating...")
exc_save_magn = path + "\\" + Excel_name + ".xlsx"

while(True):
    try:
        if os.access(exc_save_magn,os.F_OK):
            print("modifying...")
            book = xl.load_workbook(exc_save_magn)

            sheet1 = book["Overtime"]
            sheet1.append([DEP_f])
            sheet1.append(samp_STD_list)
            sheet1.append(samp_ave_list)

            sheet2 = book["Estimation"]
            sheet2.append([DEP_f, shift_comp_area_mean, shift_noComp_area_mean, samp_temp, samp_area_mean_unmodify, temp_gap, temp_compen_mean])

            sheet3 = book["Curve_fitting"]
            parameter = parameter.tolist()
            sheet3.append(parameter)
            sheet3.append([DEP_f, "r_square"])
            sheet3.append([r_square])

            sheet4 = book["Estimation_respective_Comp"]
            shift_comp_modi = shift_comp_modi.tolist()
            sheet4.append([DEP_f])
            sheet4.append(shift_comp_modi)

            sheet5 = book["Estimation_respective_noComp"]
            shift_noComp_modi = shift_noComp_modi.tolist()
            sheet5.append([DEP_f])
            sheet5.append(shift_noComp_modi)


        else:
            print("creating...")
            book = xl.Workbook()

            sheet1 = book.create_sheet("Overtime")
            sheet1.append([DEP_f])
            sheet1.append(samp_STD_list)
            sheet1.append(samp_ave_list)


            sheet2 = book.create_sheet("Estimation")
            sheet2.cell(row=1, column=1).value = 'DEP Frequency'
            sheet2.cell(row=1, column=2).value = 'Frequency Shift Modified[MHz]'
            sheet2.cell(row=1, column=3).value = 'Frequency Shift UnCompensated[MHz]'
            sheet2.cell(row=1, column=4).value = 'Temperature Speculation[℃]'
            sheet2.cell(row=1, column=5).value = 'Frequency Shift UnModified[MHz]'
            sheet2.cell(row=1, column=6).value = 'Temperature Gap[℃]'
            sheet2.cell(row=1, column=7).value = 'Temperature Compensation Average[MHz]'
            sheet2.append([DEP_f, shift_comp_area_mean, shift_noComp_area_mean, samp_temp, samp_area_mean_unmodify, temp_gap, temp_compen_mean])

            sheet3 = book.create_sheet("Curve_fitting")
            sheet3.cell(row=1, column=1).value = 'C0'
            sheet3.cell(row=1, column=2).value = 'C1'
            sheet3.cell(row=1, column=3).value = 'k'
            parameter = parameter.tolist()
            sheet3.append(parameter)
            sheet3.append([DEP_f, "r_square"])
            sheet3.append([r_square])

            sheet4 = book.create_sheet("Estimation_respective_Comp")
            shift_comp_modi = shift_comp_modi.tolist()
            sheet4.append([DEP_f])
            sheet4.append(shift_comp_modi)

            sheet5 = book.create_sheet("Estimation_respective_noComp")
            shift_noComp_modi = shift_noComp_modi.tolist()
            sheet5.append([DEP_f])
            sheet5.append(shift_noComp_modi)

        book.save(exc_save_magn)
        print("Saved")
        break
    except:
        print("Falied!!! Close the file")
        time.sleep(1)

print("Finished 1")



#=====matrix_form======
matrix_noComp=np.zeros((62, 24))
for i in range(0, 62):
    for j in range(0, 24):
        matrix_noComp[i][j] = shift_noComp_modi[1487 - (23 - j) - 24 * i]  ###########
exc_save_magn = path + "\\" + DEP_f + ".xlsx"
print("creating...")
book2 = xl.Workbook()

sheet6 = book2.worksheets[0]
for j in range(0, 24):
    matrix_noComp_list = matrix_noComp[:,j].tolist()
    sheet6.append(matrix_noComp_list)

book2.save(exc_save_magn)
print("Finished")


